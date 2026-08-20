"""
import_history_hs_code_importer.py — SKU/OEM 수입이력 HS코드 리서치 결과 Excel →
import_history.hs_code / hs_code_confidence 일괄 반영.

hs_code_importer.py(메인테이블용)와 같은 패턴이지만, 매칭 키가 다르다:
메인테이블은 (product_type, retailer, rank) 3중 키지만, 여기(import_history)는
SKU명(sku_name) 단일 키 — PATCH /api/import-history/hs-code와 동일한 단위.

원본 컬럼(Results 시트): SKU명, 원본행수, 영문상품명, MC, HS코드, 신뢰도, 판정상태,
판정근거, 근거

confidence 등급별 처리:
  - high        : hs_code 그대로 반영
  - medium      : hs_code 반영하되 hs_code_confidence='medium'으로 저장
  - low/very_low/기타: 반영하지 않음 (의도적 skip)

같은 SKU명이 DB에 여러 행(수입이력 레코드) 존재할 수 있어 UPDATE ... WHERE
sku_name = :sku_name 로 일괄 반영한다 — PATCH 엔드포인트와 동일한 단위.
"""
from __future__ import annotations
from functools import lru_cache
from io import BytesIO
from sqlalchemy.ext.asyncio import AsyncSession
from sqlalchemy import text

_EXPECTED_HEADER_PREFIX = ("SKU명", "원본행수")
_ALLOWED_CONFIDENCE = {"high", "medium"}
_MAX_UNMATCHED_SAMPLES = 200


async def import_history_hs_codes(content: bytes, db: AsyncSession) -> dict:
    openpyxl_mod = _openpyxl()
    wb = openpyxl_mod.load_workbook(BytesIO(content), read_only=True)

    updated = 0
    skipped_low_confidence = 0
    skipped_unmatched = 0
    total = 0
    unmatched_samples: list[dict] = []

    def _record_unmatched(reason: str, sku_name: str, hs_code: str):
        nonlocal skipped_unmatched
        skipped_unmatched += 1
        if len(unmatched_samples) < _MAX_UNMATCHED_SAMPLES:
            unmatched_samples.append({"sku_name": sku_name, "hs_code": hs_code, "reason": reason})

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        rows = ws.iter_rows(values_only=True)
        header = next(rows, None)
        if not header or tuple(header[:2]) != _EXPECTED_HEADER_PREFIX:
            continue

        for r in rows:
            if not r or not r[0] or not r[4]:  # SKU명 또는 HS코드 없으면 집계 대상도 아님
                continue
            total += 1

            sku_name = str(r[0]).strip()
            hs_code = str(r[4]).strip()
            confidence = str(r[5]).strip().lower() if r[5] else None

            if confidence not in _ALLOWED_CONFIDENCE:
                skipped_low_confidence += 1
                continue

            result = await db.execute(text("""
                UPDATE import_history
                SET hs_code = :hs_code, hs_code_confidence = :confidence
                WHERE sku_name = :sku_name
            """), {"hs_code": hs_code, "confidence": confidence, "sku_name": sku_name})
            if result.rowcount:
                updated += result.rowcount
            else:
                _record_unmatched("no_matching_db_row", sku_name, hs_code)

    await db.commit()
    return {
        "total_rows": total,
        "updated": updated,
        "skipped_low_confidence": skipped_low_confidence,
        "skipped_unmatched": skipped_unmatched,
        "unmatched_samples": unmatched_samples,
    }


@lru_cache(maxsize=1)
def _openpyxl():
    import openpyxl
    return openpyxl
