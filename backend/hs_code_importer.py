"""
hs_code_importer.py — 상품별(유형×유통사×순위) HS코드 리서치 결과 Excel →
product_sourcing_item.hs_code / hs_code_confidence 일괄 반영.

원본 컬럼: 유형, 유통사(서브타이틀 문자열 — product_sourcing_importer.py의
_parse_subtitle과 동일 포맷), rank_col, 브랜드(원본), 영어상품명, 판정,
판정근거, hs_code, confidence, reason, product_identity_note, status

품목유형 단위가 아니라 (product_type, retailer, rank) — 즉 product_sourcing_item의
UniqueConstraint(product_type, retailer, rank)와 정확히 같은 단위로 매칭해
UPDATE한다. 유형/유통사/순위가 기존 DB에 없는 행(오탈자, 리서치 시점 차이 등)은
매칭 실패로 세어 skipped_unmatched에 넣고 계속 진행한다.

confidence 등급별 처리:
  - high        : hs_code 그대로 반영
  - medium      : hs_code 반영하되 hs_code_confidence='medium'으로 저장 —
                   프론트에서 "(검토 필요)" 표시에 사용
  - low/very_low/기타: 아예 반영하지 않음 (의도적 skip, skipped_low_confidence로 별도 집계)

이 파일은 앞으로도 반복해서 새 리서치 결과가 올라올 예정이라, 실패한 행이
"왜" 실패했는지(파싱 실패/랭크 오류/DB에 매칭되는 행 없음)를 매번 사람이
DB를 뒤져서 확인할 필요 없이 응답에 그대로 담아 돌려준다 — unmatched_samples 참고.
"""
from __future__ import annotations
from functools import lru_cache
from io import BytesIO
from sqlalchemy.ext.asyncio import AsyncSession
from sqlalchemy import text

from product_sourcing_importer import _parse_subtitle

_EXPECTED_HEADER_PREFIX = ("유형", "유통사", "rank_col")
_ALLOWED_CONFIDENCE = {"high", "medium"}  # low/very_low는 업로드 대상에서 제외
_MAX_UNMATCHED_SAMPLES = 200  # 응답이 무한정 커지지 않게 상한


async def import_hs_codes(content: bytes, db: AsyncSession) -> dict:
    openpyxl_mod = _openpyxl()
    wb = openpyxl_mod.load_workbook(BytesIO(content), read_only=True)

    updated = 0
    skipped_low_confidence = 0
    skipped_unmatched = 0
    total = 0
    unmatched_samples: list[dict] = []

    def _record_unmatched(reason: str, product_type: str, retailer_raw: str, rank, hs_code: str):
        nonlocal skipped_unmatched
        skipped_unmatched += 1
        if len(unmatched_samples) < _MAX_UNMATCHED_SAMPLES:
            unmatched_samples.append({
                "product_type": product_type, "retailer_raw": retailer_raw,
                "rank": rank, "hs_code": hs_code, "reason": reason,
            })

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        rows = ws.iter_rows(values_only=True)
        header = next(rows, None)
        if not header or tuple(header[:3]) != _EXPECTED_HEADER_PREFIX:
            continue

        for r in rows:
            if not r or not r[0] or not r[7]:  # 유형 또는 hs_code 없으면 집계 대상도 아님
                continue
            total += 1

            product_type = str(r[0]).strip()
            retailer_raw = str(r[1] or "").strip()
            subtitle = _parse_subtitle(retailer_raw)
            rank = r[2]
            hs_code = str(r[7]).strip()
            confidence = str(r[8]).strip().lower() if r[8] else None

            if confidence not in _ALLOWED_CONFIDENCE:
                skipped_low_confidence += 1  # low/very_low/미상 — 검토 없이는 반영하지 않음, 실패 아님
                continue

            if not subtitle:
                _record_unmatched("retailer_text_parse_failed", product_type, retailer_raw, rank, hs_code)
                continue
            if rank is None:
                _record_unmatched("rank_missing", product_type, retailer_raw, rank, hs_code)
                continue

            try:
                rank_int = int(rank)
            except (TypeError, ValueError):
                _record_unmatched("rank_not_integer", product_type, retailer_raw, rank, hs_code)
                continue

            result = await db.execute(text("""
                UPDATE product_sourcing_item
                SET hs_code = :hs_code, hs_code_confidence = :confidence
                WHERE product_type = :pt AND retailer = :retailer AND rank = :rank
            """), {
                "hs_code": hs_code, "confidence": confidence,
                "pt": product_type, "retailer": subtitle["retailer"], "rank": rank_int,
            })
            if result.rowcount:
                updated += result.rowcount
            else:
                _record_unmatched("no_matching_db_row", product_type, retailer_raw, rank, hs_code)

    await db.commit()
    return {
        "total_rows": total,
        "updated": updated,
        "skipped_low_confidence": skipped_low_confidence,
        "skipped_unmatched": skipped_unmatched,
        "unmatched_samples": unmatched_samples,
    }


@lru_cache(maxsize=1)
def _openpyxl():
    import openpyxl
    return openpyxl
