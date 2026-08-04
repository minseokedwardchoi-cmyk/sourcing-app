"""
product_sourcing_importer.py — 품목 소싱 리서치 Excel → product_sourcing_item 적재

원본 파일 구조 (ESI Top40 리서치, 확인 완료):
  시트 "유형별카드": 품목(유형)별 카드 레이아웃. "■ 품목명" 행으로 품목 블록이
    시작하고, 그 안에 "{유통사}  ·  {method}   (상위 N개 · PB 제외)" 형태의
    행으로 유통사 서브블록이 시작한다. 서브블록 안에는 브랜드(한국명)/가격(USD)/
    원산지/단량/핵심기준(...)/병행수입가능여부/리콜 이력/품질·표시/법적·평판
    리스크/5년내 이슈/비고 등의 "라벨 행"이 순위(1~N위) 순서로 값을 담고 있다.
    이미지는 셀 값이 아니라 삽입된 이미지 객체라 텍스트로 읽을 수 없다.
  시트 "상품리스트(raw)": 유형/유통사/순위 1행 = 상품 1개. 브랜드(영문), 상품명,
    URL, 이미지URL, 평점, 리뷰수, 실측여부를 담고 있고 리스크 판정 컬럼은 없다.

두 시트를 (유형, 유통사, 순위)로 조인해야 완전한 레코드가 나온다. 유통사는
사용자가 요청한 4곳(월마트/아마존/샘스클럽/이온몰)만 적재한다 — 원본에는
Al Campo/Conad/Cosi Comodo 등 유럽 유통사 데이터도 섞여 있지만 이번 페이지의
범위 밖이라 건너뛴다.
"""
from __future__ import annotations
import re
from functools import lru_cache
from io import BytesIO
from sqlalchemy.ext.asyncio import AsyncSession
from sqlalchemy import text
from models import ProductSourcingItem


@lru_cache(maxsize=1)
def _openpyxl():
    import openpyxl
    return openpyxl


RETAILER_PREFIX_MAP: dict[str, str] = {
    "월마트": "walmart",
    "아마존": "amazon",
    "샘스클럽": "samsclub",
    "이온몰": "aeon",
}

_SUBTITLE_RE = re.compile(
    r"^(?P<retailer_kr>\S+)\s*(?:·\s*(?P<method>\S*)\s*)?\(상위\s*(?P<n>\d+)개"
)

# "유형별카드" 라벨 행 → 내부 필드명. 라벨은 개행이 섞여 있어 먼저 개행을 지우고 비교한다.
_EXACT_LABEL_MAP: dict[str, str] = {
    "브랜드(한국명)": "brand_kr",
    "가격(USD)": "price_usd",
    "원산지": "origin",
    "단량": "unit",
    "병행수입가능여부": "parallel_import",
    "리콜 이력": "recall_status",
    "품질·표시": "quality_label_status",
    "법적·평판 리스크": "legal_risk_status",
    "5년내 이슈": "five_year_issue",
    "영어(원어) 상품명": "product_name_en_card",
}


def _clean_label(raw: str) -> str:
    return re.sub(r"\s+", "", raw.replace("\n", ""))


# 위 _EXACT_LABEL_MAP은 개행 제거 후 비교하므로 공백도 통일해서 다시 만든다.
_LABEL_MAP: dict[str, str] = {_clean_label(k): v for k, v in _EXACT_LABEL_MAP.items()}


def _classify_label(raw: str) -> tuple[str, str | None]:
    """라벨 문자열 → (내부필드명, 부가정보). '핵심기준'/'비고' 계열은 품목마다
    괄호 안 텍스트가 달라서(산도/난황 함량/냉압착·정제여부 등) 접두사로 잡는다."""
    cleaned = _clean_label(raw)
    if cleaned in _LABEL_MAP:
        return _LABEL_MAP[cleaned], None
    if cleaned.startswith("핵심기준"):
        m = re.search(r"\((.*)\)", raw.replace("\n", ""))
        return "key_criteria", (m.group(1) if m else None)
    if cleaned.startswith("비고"):
        return "notes", None
    return "unknown", None


def _parse_subtitle(text: str) -> dict | None:
    m = _SUBTITLE_RE.match(text.strip())
    if not m:
        return None
    retailer_kr = m.group("retailer_kr")
    canonical = RETAILER_PREFIX_MAP.get(retailer_kr)
    if not canonical:
        return None
    return {
        "retailer": canonical,
        "retailer_label": retailer_kr,
        "ranking_method": (m.group("method") or "").strip() or None,
        "sample_note": text.strip(),
        "_n": int(m.group("n")),
    }


def _to_float(val):
    if val is None:
        return None
    try:
        return float(val)
    except (TypeError, ValueError):
        return None


def _to_int(val):
    f = _to_float(val)
    return int(f) if f is not None else None


def _clean_url(val):
    """일부 유통사(아마존/이온몰) 행은 URL 칸이 비어 있고, 이미지URL 칸에 리서치
    작업자의 로컬 캐시 경로(C:\\Users\\...)가 남아 있다. 웹에서 못 쓰는 값이라
    http(s) 링크가 아니면 버린다."""
    if not val or not isinstance(val, str) or not val.startswith(("http://", "https://")):
        return None
    return val


def parse_card_sheet(ws) -> list[dict]:
    """'유형별카드' 시트를 파싱해 (product_type, retailer, rank) 단위 레코드로 펼친다.

    각 레코드에 "_image_row"(그 유통사 서브블록의 '이미지' 라벨 행, 0-index)를
    함께 담아둔다 — 셀 값이 아니라 삽입된 그림 객체라 여기서는 위치만 기록하고,
    실제 그림 바이트는 build_records에서 extract_card_images()로 뽑아 이 좌표로
    매칭한다.
    """
    records: list[dict] = []

    current_type: str | None = None
    current_retailer: dict | None = None
    current_data: dict[str, list] = {}
    current_key_criteria_label: str | None = None
    current_image_row: int | None = None

    def flush():
        if not current_type or not current_retailer:
            return
        n = current_retailer["_n"]
        retailer_fields = {k: v for k, v in current_retailer.items() if k != "_n"}
        for i in range(n):
            def at(field):
                vals = current_data.get(field)
                if not vals or i >= len(vals):
                    return None
                v = vals[i]
                return v if v not in (None, "") else None

            if at("brand_kr") is None:
                continue

            records.append({
                "product_type": current_type,
                "rank": i + 1,
                **retailer_fields,
                "brand_kr": at("brand_kr"),
                "price_usd": _to_float(at("price_usd")),
                "origin": at("origin"),
                "unit": at("unit"),
                "key_criteria_label": current_key_criteria_label,
                "key_criteria_value": at("key_criteria"),
                "parallel_import": at("parallel_import"),
                "recall_status": at("recall_status"),
                "quality_label_status": at("quality_label_status"),
                "legal_risk_status": at("legal_risk_status"),
                "five_year_issue": at("five_year_issue"),
                "notes": at("notes"),
                "product_name_en_card": at("product_name_en_card"),
                "_image_row": current_image_row,
                "_image_col": i + 1,
            })

    for row_idx, row in enumerate(ws.iter_rows(values_only=True)):
        first = row[0]
        if not isinstance(first, str) or not first.strip():
            continue
        text = first.strip()

        if text.startswith("■"):
            flush()
            current_type = text.lstrip("■").strip()
            current_retailer = None
            current_data = {}
            current_key_criteria_label = None
            current_image_row = None
            continue

        if "(상위" in text:
            parsed = _parse_subtitle(text)
            flush()
            current_retailer = parsed  # None이면(대상 외 유통사) 이후 라벨 행은 무시됨
            current_data = {}
            current_key_criteria_label = None
            current_image_row = None
            continue

        if text.startswith("항목"):
            continue

        if text == "이미지":
            current_image_row = row_idx
            continue

        field, extra = _classify_label(text)
        if field == "unknown" or current_retailer is None:
            continue
        if field == "key_criteria" and extra:
            current_key_criteria_label = extra
        current_data[field] = list(row[1:])

    flush()
    return records


_MIME_BY_EXT = {
    "jpg": "image/jpeg", "jpeg": "image/jpeg", "png": "image/png",
    "gif": "image/gif", "bmp": "image/bmp", "webp": "image/webp",
}


def extract_card_images(ws) -> dict[tuple[int, int], tuple[bytes, str]]:
    """'유형별카드' 시트에 삽입된 그림들을 (행, 열) 좌표 → (바이트, MIME) 로 인덱싱.

    셀 값이 아니라 워크시트에 얹힌 그림 객체(ws._images)라서, 각 그림의 앵커
    (anchor.row/col, 0-index)로 어느 순위 칸 위에 놓여있는지 알아낸다.
    """
    out: dict[tuple[int, int], tuple[bytes, str]] = {}
    for img in getattr(ws, "_images", []):
        anchor = getattr(img, "anchor", None)
        frm = getattr(anchor, "_from", None)
        if frm is None:
            continue
        ext = (img.path or "").rsplit(".", 1)[-1].lower()
        mime = _MIME_BY_EXT.get(ext, "image/jpeg")
        try:
            data = img._data()
        except Exception:
            continue
        out[(frm.row, frm.col)] = (data, mime)
    return out


def parse_raw_sheet(ws) -> dict[tuple[str, str, str], list[dict]]:
    """'상품리스트(raw)' 시트를 (유형, 유통사, 브랜드) → 상세정보 리스트로 인덱싱.

    카드 시트의 순위는 "PB 제외"로 다시 매겨진 값이라 raw 시트의 원본 순위와
    그대로 대응하지 않는다(카드가 raw의 특정 NB 항목을 더 건너뛰는 경우도 있어
    PB만 걸러내는 걸로도 못 맞는다). 대신 두 시트에 공통으로 존재하는
    (유형, 유통사, 브랜드) — 카드의 "브랜드(한국명)"과 raw의 "브랜드"는 표기가
    동일함 — 로 묶어두고, 병합 시 가격까지 맞는 항목을 우선으로 소비한다.
    """
    header = [str(c).strip() if c is not None else "" for c in next(ws.iter_rows(min_row=1, max_row=1, values_only=True))]
    idx = {name: i for i, name in enumerate(header)}

    out: dict[tuple[str, str, str], list[dict]] = {}
    for row in ws.iter_rows(min_row=2, values_only=True):
        product_type = row[idx.get("유형", 0)]
        retailer = row[idx.get("유통사", 1)]
        brand = row[idx["브랜드"]] if "브랜드" in idx else None
        if not product_type or not retailer or not brand:
            continue
        if retailer not in RETAILER_PREFIX_MAP.values():
            continue
        key = (str(product_type).strip(), str(retailer).strip(), str(brand).strip())
        out.setdefault(key, []).append({
            "price_usd": _to_float(row[idx["가격(USD)"]]) if "가격(USD)" in idx else None,
            "brand_en": row[idx["브랜드(영문)"]] if "브랜드(영문)" in idx else None,
            "product_name_en": row[idx["상품명"]] if "상품명" in idx else None,
            "rating": _to_float(row[idx["평점"]]) if "평점" in idx else None,
            "review_count": _to_int(row[idx["리뷰수"]]) if "리뷰수" in idx else None,
            "url": _clean_url(row[idx["URL"]]) if "URL" in idx else None,
            "image_url": _clean_url(row[idx["이미지URL"]]) if "이미지URL" in idx else None,
            "verified_flag": row[idx["실측여부"]] if "실측여부" in idx else None,
        })
    return out


def build_records(file_bytes: bytes) -> list[dict]:
    openpyxl = _openpyxl()
    # 이미지(삽입된 그림 객체)를 읽으려면 read_only가 아닌 일반 모드로 열어야 한다.
    # 파일 하나를 두 번 여는 대신, 이 한 번의 로드를 카드/raw 시트 파싱에 모두 쓴다.
    wb = openpyxl.load_workbook(BytesIO(file_bytes), data_only=True)
    card_ws = wb["유형별카드"]
    raw_ws = wb["상품리스트(raw)"]

    card_records = parse_card_sheet(card_ws)
    card_images = extract_card_images(card_ws)
    raw_index = parse_raw_sheet(raw_ws)

    merged: list[dict] = []
    for rec in card_records:
        candidates = raw_index.get((rec["product_type"], rec["retailer"], rec["brand_kr"] or "")) or []
        extra = None
        # 1) 가격까지 일치하는 미사용 항목 우선
        for cand in candidates:
            if cand.get("_used"):
                continue
            if rec["price_usd"] is not None and cand["price_usd"] is not None and abs(cand["price_usd"] - rec["price_usd"]) < 0.005:
                extra = cand
                break
        # 2) 없으면 브랜드만 맞는 미사용 항목 순서대로
        if extra is None:
            for cand in candidates:
                if not cand.get("_used"):
                    extra = cand
                    break
        if extra is not None:
            extra["_used"] = True
        extra = extra or {}

        product_name_en = extra.get("product_name_en") or rec.get("product_name_en_card")
        image_url = extra.get("image_url")
        image_data = image_mime = None
        # raw 시트에 실제 호스팅된 이미지 URL이 없을 때만(아마존/이온몰), 카드
        # 시트에 삽입된 그림을 대신 쓴다 — 있는 URL을 그림으로 덮어쓰지 않는다.
        if not image_url:
            img_row, img_col = rec.get("_image_row"), rec.get("_image_col")
            if img_row is not None:
                found = card_images.get((img_row, img_col))
                if found:
                    image_data, image_mime = found
        rec.pop("product_name_en_card", None)
        rec.pop("_image_row", None)
        rec.pop("_image_col", None)
        merged.append({
            **rec,
            "product_name_en": product_name_en,
            "brand_en": extra.get("brand_en"),
            "rating": extra.get("rating"),
            "review_count": extra.get("review_count"),
            "url": extra.get("url"),
            "image_url": image_url,
            "image_data": image_data,
            "image_mime": image_mime,
            "verified_flag": extra.get("verified_flag"),
        })
    return merged


async def _ensure_image_columns(db: AsyncSession) -> None:
    """image_data/image_mime 컬럼이 없으면 추가한다 (기존 배포에 테이블만 있고
    이 두 컬럼은 없는 경우 대비). main.py의 startup 마이그레이션이 다른 ALTER
    문과 한 트랜잭션에 묶여있어 그중 하나라도 실패하면 뒤엣것까지 조용히
    무시될 수 있어서, 업로드 시점에 한 번 더 독립적으로 보장해둔다."""
    for col_sql in [
        "ALTER TABLE product_sourcing_item ADD COLUMN IF NOT EXISTS image_data BYTEA",
        "ALTER TABLE product_sourcing_item ADD COLUMN IF NOT EXISTS image_mime VARCHAR(50)",
    ]:
        try:
            await db.execute(text(col_sql))
        except Exception:
            await db.rollback()


async def import_product_sourcing(file_bytes: bytes, db: AsyncSession) -> dict:
    """Excel(카드 시트 + raw 시트) → product_sourcing_item 테이블 전체 재적재.

    품목/유통사 구성이 매번 파일 전체를 대체하는 성격의 리서치 자료라, 증분
    업데이트 대신 매 업로드마다 테이블을 비우고 새로 채운다(재수입해도 중복이
    쌓이지 않도록).
    """
    records = build_records(file_bytes)

    await _ensure_image_columns(db)
    await db.execute(ProductSourcingItem.__table__.delete())

    CHUNK = 2000
    for i in range(0, len(records), CHUNK):
        await db.execute(ProductSourcingItem.__table__.insert(), records[i:i + CHUNK])
        await db.flush()

    await db.commit()

    types = sorted({r["product_type"] for r in records})
    return {
        "inserted": len(records),
        "product_type_count": len(types),
    }
