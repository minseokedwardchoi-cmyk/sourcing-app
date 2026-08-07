"""
tariff_rate_importer.py — 관세청_품목번호별 관세율표(data.go.kr) Excel → tariff_rate 적재

원본 컬럼 순서(고정): 품목번호, 관세율구분, 관세율, 단위당세액, 기준가격,
적용국가구분, 용도세율구분, 적용개시일(YYYYMMDD), 적용만료일(YYYYMMDD)

파일에 시트가 여러 개(예: '1.1', '2.12') 있는데 실제로 내용이 거의 동일한
버전/차수 복사본이다 — 처음 발견되는 형식 일치 시트 하나만 사용하고
나머지는 건너뛴다 (전부 합쳐서 파싱하면 처리량이 불필요하게 두 배가 됨).
매번 전체 재적재(TRUNCATE 후 INSERT) 방식 — 관세율은 개정되면 옛 값이
남아있으면 안 되므로.

중요: openpyxl로 38만 행을 파싱하는 건 완전히 동기적(sync)인 CPU 작업이라,
이걸 그냥 async 함수 안에서 돌리면 그 몇 초~몇십 초 동안 이벤트 루프가
막혀서 서버가 Render의 health check(5초 타임아웃)에 응답을 못 하고
"unhealthy"로 판정돼 강제 재시작당한다 (2026-08 실제로 확인된 원인 —
OOM이 아니었음). 그래서 파싱 부분만 asyncio.to_thread로 별도 스레드에
돌려서 메인 이벤트 루프가 막히지 않게 한다.
"""
from __future__ import annotations
import asyncio
import re
from functools import lru_cache
from io import BytesIO
from sqlalchemy.ext.asyncio import AsyncSession
from sqlalchemy import text

_EXPECTED_HEADER_PREFIX = ("품목번호", "관세율구분", "관세율")
# 배치를 작게 유지하면 배치 사이 await 지점이 자주 생겨 이벤트 루프가
# health check 등 다른 요청을 처리할 기회가 늘어난다.
_BATCH_SIZE = 5000


@lru_cache(maxsize=1)
def _openpyxl():
    import openpyxl
    return openpyxl


def _parse_date(raw) -> str | None:
    if raw is None:
        return None
    s = str(raw).strip()
    if len(s) != 8 or not s.isdigit():
        return None
    return f"{s[0:4]}-{s[4:6]}-{s[6:8]}"


def _parse_int(raw) -> int | None:
    if raw is None or raw == "":
        return None
    try:
        return int(raw)
    except (TypeError, ValueError):
        return None


def _parse_tariff_records(content: bytes) -> list[dict]:
    """동기(sync) 파싱 함수 — asyncio.to_thread로 별도 스레드에서 호출된다.
    이벤트 루프를 막지 않으려고 DB 작업과 분리해뒀다."""
    openpyxl = _openpyxl()
    wb = openpyxl.load_workbook(BytesIO(content), read_only=True)

    dedup: dict[tuple, dict] = {}
    used_sheet = False
    for sheet_name in wb.sheetnames:
        if used_sheet:
            break  # 형식이 맞는 시트를 이미 하나 처리했으면 나머지(중복 사본)는 건너뜀
        ws = wb[sheet_name]
        rows = ws.iter_rows(values_only=True)
        header = next(rows, None)
        if not header or tuple(header[:3]) != _EXPECTED_HEADER_PREFIX:
            continue  # 예상 형식이 아닌 시트는 건너뜀 (표지/설명 시트 등)
        used_sheet = True

        for r in rows:
            if not r or not r[0] or r[1] is None or r[2] is None:
                continue
            # 상품 쪽(product_sourcing_item.hs_code)은 '2008.11-1000'처럼 점/대시가
            # 섞인 표기라, 여기서도 숫자만 남겨 저장해둬야 main.py의
            # _attach_cost_estimates가 인덱스를 그대로 타는 단순 등가비교로
            # 매칭할 수 있다 (조회 시점에 함수로 정규화하면 인덱스를 못 타 느려짐).
            hs_code   = re.sub(r"\D", "", str(r[0]).strip())
            if not hs_code:
                continue
            rate_type = str(r[1]).strip()
            rate_pct  = float(r[2])
            country_group = _parse_int(r[5]) if len(r) > 5 else None
            eff_from  = _parse_date(r[7]) if len(r) > 7 else None
            eff_to    = _parse_date(r[8]) if len(r) > 8 else None

            key = (hs_code, rate_type, eff_from)
            dedup[key] = {
                "hs_code": hs_code, "rate_type": rate_type, "rate_pct": rate_pct,
                "country_group": country_group, "eff_from": eff_from, "eff_to": eff_to,
            }

    return list(dedup.values())


async def import_tariff_rates(content: bytes, db: AsyncSession) -> dict:
    records = await asyncio.to_thread(_parse_tariff_records, content)

    await db.execute(text("TRUNCATE TABLE tariff_rate"))
    await db.commit()

    # 예전에는 배치당 (:hs_code_0, :rate_type_0, ...), (:hs_code_1, ...) 식으로
    # 행마다 이름 있는 파라미터를 새로 만들어 하나의 VALUES 문자열로 합쳤는데,
    # 배치가 커지면(예: 1000행 = 파라미터 6000개) SQLAlchemy의 text() 바인드
    # 파라미터 자동 인식이 일부 이름을 잘못 파싱해서 최종 SQL에 치환 안 된
    # ":xxx_123" 리터럴 콜론이 그대로 남아 "syntax error at or near ':'"로
    # 터지거나(제조사의 다른 함수 _recompute_and_store_cost_estimates에서 실제
    # 재현/확인됨) 조용히 응답 없이 멎는 원인이 됐다. unnest()로 배열을 통째로
    # 바인딩하면 배치 크기와 무관하게 파라미터가 항상 6개뿐이라 이 문제 자체가
    # 발생할 수 없다.
    total_inserted = 0
    hs_codes_seen: set[str] = set()
    for i in range(0, len(records), _BATCH_SIZE):
        chunk = records[i:i + _BATCH_SIZE]
        await db.execute(text("""
            INSERT INTO tariff_rate (hs_code, rate_type, rate_pct, applies_country_group, effective_from, effective_to)
            SELECT * FROM unnest(
                :hs_codes::varchar[], :rate_types::varchar[], :rate_pcts::numeric[],
                :country_groups::integer[], :eff_froms::date[], :eff_tos::date[]
            )
        """), {
            "hs_codes": [rec["hs_code"] for rec in chunk],
            "rate_types": [rec["rate_type"] for rec in chunk],
            "rate_pcts": [rec["rate_pct"] for rec in chunk],
            "country_groups": [rec["country_group"] for rec in chunk],
            "eff_froms": [rec["eff_from"] for rec in chunk],
            "eff_tos": [rec["eff_to"] for rec in chunk],
        })
        await db.commit()
        total_inserted += len(chunk)
        hs_codes_seen.update(rec["hs_code"] for rec in chunk)

    return {
        "inserted": total_inserted,
        "hs_code_count": len(hs_codes_seen),
    }
