"""
product_sourcing_exporter.py — product_sourcing_item 테이블 → 원본 엑셀
("유형별카드" 시트) 형식의 .xlsx 재구성.

product_sourcing_importer.py의 역방향 작업. id 순서가 원본 파일에서
파싱된 순서(품목 블록 → 유통사 서브블록 → 순위) 그대로 보존되므로,
그 순서만 따라가며 블록 경계(품목/유통사·sample_note 변경 시점)를
다시 나누면 원본과 동일한 레이아웃을 복원할 수 있다.

"항목" 라벨 행의 나머지 칸(1위/2위/…N위)은 원본 파싱 시 값을 읽지
않고 건너뛰어서(문자열이라 텍스트로 안 읽음) DB에 남아있지 않다 —
여기서는 rank 값으로부터 다시 합성한다.
"""
from __future__ import annotations
from io import BytesIO
from typing import Any, Mapping, Sequence

from openpyxl import Workbook
from openpyxl.drawing.image import Image as XLImage
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

_LABEL_FIELDS: list[tuple[str, str | None]] = [
    ("브랜드(한국명)", "brand_kr"),
    ("가격(USD)", "price_usd"),
    ("원산지", "origin"),
    ("단량", "unit"),
    ("__key_criteria__", None),
    ("병행수입가능여부", "parallel_import"),
    ("리콜 이력", "recall_status"),
    ("품질·표시", "quality_label_status"),
    ("법적·평판 리스크", "legal_risk_status"),
    ("5년내 이슈", "five_year_issue"),
    ("비고", "notes"),
    ("영어(원어) 상품명", "product_name_en"),
]

_TYPE_FILL = PatternFill("solid", fgColor="1F2937")
_TYPE_FONT = Font(bold=True, color="FFFFFF", size=13)
_SUB_FILL = PatternFill("solid", fgColor="DBEAFE")
_SUB_FONT = Font(bold=True, color="1E3A8A", size=11)
_ITEM_FILL = PatternFill("solid", fgColor="F3F4F6")
_ITEM_FONT = Font(bold=True, size=10)
_LABEL_FONT = Font(bold=True, size=10)
_WRAP = Alignment(vertical="center", wrap_text=True)

_IMG_MAX_PX = 90
_IMG_ROW_HEIGHT = 70
_VALUE_COL_WIDTH = 15
_LABEL_COL_WIDTH = 22


def _group_by_block(rows: Sequence[Mapping[str, Any]]):
    """id 순서(=원본 파일 순서) rows를
    [(product_type, [{"retailer_label","sample_note","rows":[...]}, ...]), ...] 로 묶는다."""
    types: list[tuple[str, list[dict]]] = []
    cur_type = object()
    cur_sub_key = object()
    subblocks: list[dict] | None = None
    cur_sub: dict | None = None

    for row in rows:
        if row["product_type"] != cur_type:
            cur_type = row["product_type"]
            subblocks = []
            types.append((cur_type, subblocks))
            cur_sub_key = object()

        sub_key = (row["retailer_label"], row["sample_note"])
        if sub_key != cur_sub_key:
            cur_sub_key = sub_key
            cur_sub = {"retailer_label": row["retailer_label"], "sample_note": row["sample_note"], "rows": []}
            subblocks.append(cur_sub)

        cur_sub["rows"].append(row)

    return types


def _fill_row(ws, row_idx: int, last_col: int, fill: PatternFill) -> None:
    for col in range(1, last_col + 1):
        ws.cell(row=row_idx, column=col).fill = fill


def _num(v):
    return float(v) if v is not None else None


def build_original_format_workbook(rows: Sequence[Mapping[str, Any]]) -> BytesIO:
    wb = Workbook()
    ws = wb.active
    ws.title = "유형별카드"

    r = 1
    max_col_seen = 1

    for product_type, subblocks in _group_by_block(rows):
        block_max_rank = max((row["rank"] for sb in subblocks for row in sb["rows"]), default=1)
        last_col = block_max_rank + 1
        max_col_seen = max(max_col_seen, last_col)

        c = ws.cell(row=r, column=1, value=f"■ {product_type}")
        c.font = _TYPE_FONT
        _fill_row(ws, r, last_col, _TYPE_FILL)
        ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=last_col)
        r += 1

        for sb in subblocks:
            sub_rows = sb["rows"]
            sub_max_rank = max(row["rank"] for row in sub_rows)
            sub_last_col = sub_max_rank + 1

            header_text = sb["sample_note"] or f'{sb["retailer_label"] or ""} (상위 {len(sub_rows)}개)'
            c = ws.cell(row=r, column=1, value=header_text)
            c.font = _SUB_FONT
            _fill_row(ws, r, sub_last_col, _SUB_FILL)
            ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=sub_last_col)
            r += 1

            item_row = r
            ws.cell(row=item_row, column=1, value="항목").font = _ITEM_FONT
            _fill_row(ws, item_row, sub_last_col, _ITEM_FILL)
            for row in sub_rows:
                ws.cell(row=item_row, column=row["rank"] + 1, value=f'{row["rank"]}위').font = _ITEM_FONT
            r += 1

            key_label = next((row.get("key_criteria_label") for row in sub_rows if row.get("key_criteria_label")), None)

            for label, field in _LABEL_FIELDS:
                if label == "__key_criteria__":
                    if not key_label:
                        continue
                    label = f"핵심기준({key_label})"
                    field = "key_criteria_value"

                ws.cell(row=r, column=1, value=label).font = _LABEL_FONT
                for row in sub_rows:
                    val = row.get(field)
                    if val is None or val == "":
                        continue
                    cell = ws.cell(row=r, column=row["rank"] + 1)
                    cell.value = _num(val) if field == "price_usd" else val
                    cell.alignment = _WRAP
                r += 1

            img_row = r
            ws.cell(row=img_row, column=1, value="이미지").font = _LABEL_FONT
            ws.row_dimensions[img_row].height = _IMG_ROW_HEIGHT
            for row in sub_rows:
                data = row.get("image_data")
                if not data:
                    continue
                try:
                    img = XLImage(BytesIO(bytes(data)))
                except Exception:
                    continue
                w, h = img.width, img.height
                if w and h:
                    scale = min(_IMG_MAX_PX / w, _IMG_MAX_PX / h, 1.0)
                    img.width, img.height = w * scale, h * scale
                col_letter = get_column_letter(row["rank"] + 1)
                ws.add_image(img, f"{col_letter}{img_row}")
            r += 2  # 이미지 행 + 서브블록 간 여백

        r += 1  # 품목 블록 간 여백

    ws.column_dimensions["A"].width = _LABEL_COL_WIDTH
    for col in range(2, max_col_seen + 1):
        ws.column_dimensions[get_column_letter(col)].width = _VALUE_COL_WIDTH
    ws.freeze_panes = "B1"

    _add_flat_sheet(wb, rows)

    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf


def _add_flat_sheet(wb: Workbook, rows: Sequence[Mapping[str, Any]]) -> None:
    """원본의 '상품리스트(raw)' 시트에 대응하는 평면 데이터 시트 (이미지 없음)."""
    ws2 = wb.create_sheet("상품리스트")
    headers = [
        "유형", "유통사", "순위", "브랜드(한국명)", "브랜드(영문)", "상품명",
        "가격(USD)", "원산지", "평점", "리뷰수", "URL", "실측여부",
    ]
    ws2.append(headers)
    for cell in ws2[1]:
        cell.font = Font(bold=True)
        cell.fill = _ITEM_FILL

    for row in rows:
        ws2.append([
            row["product_type"], row.get("retailer_label"), row["rank"],
            row.get("brand_kr"), row.get("brand_en"), row.get("product_name_en"),
            _num(row.get("price_usd")), row.get("origin"), _num(row.get("rating")),
            row.get("review_count"), row.get("url"), row.get("verified_flag"),
        ])

    widths = [26, 12, 8, 20, 20, 40, 12, 16, 8, 10, 40, 12]
    for i, w in enumerate(widths, start=1):
        ws2.column_dimensions[get_column_letter(i)].width = w
    ws2.freeze_panes = "A2"
